import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.styles import Alignment
import os
import sys

# Get the path of the current application
application_path = os.path.dirname(sys.executable)

# Prompt the user to enter the month for the report
month = input("Enter the month (e.g. January, February, etc.): ")


# Read the Excel Data file and create a pivot table
df = pd.read_excel('supermarket_sales.xlsx')

df = df[['Gender', 'Product line', 'Total']]

pivot_table = df.pivot_table(index='Gender', columns='Product line', values='Total', aggfunc='sum').round(0)

# Save the pivot table to a new Excel file
pivot_table.to_excel('pivot_table.xlsx', sheet_name='Report', startrow=3)

# Load the pivot table Excel file
input_path = os.path.join(application_path, 'pivot_table.xlsx')
wb = load_workbook(input_path)
sheet = wb['Report']

# Formatting the report with a Title and Subtitle
sheet['A2'] = 'Sales Report'
sheet['A3'] = f'Month: {month}'
title = sheet['A2']
subtitle = sheet['A3']

title.font = Font('Arial', bold=True, size=20)
subtitle.font = Font('Arial', italic=True, bold=True, size=14)

# Correcting the width of the columns in the pivot table
for col_cells in sheet.columns:
    col_letter = col_cells[0].column_letter
    max_length = max(len(str(cell.value)) for cell in col_cells if cell.value is not None)
    sheet.column_dimensions[col_letter].width = max_length + 2

# Create a bar chart for the pivot table data
min_column = wb.active.min_column
max_column = wb.active.max_column
min_row = wb.active.min_row
max_row = wb.active.max_row

# Merge and center the title and subtitle 
end_column_letter = get_column_letter(max_column)
sheet.merge_cells(f'A2:{end_column_letter}2')
sheet.merge_cells(f'A3:{end_column_letter}3')
title.alignment = Alignment(horizontal='center')
subtitle.alignment = Alignment(horizontal='center')


barchart = BarChart()

data = Reference(sheet,
    min_col=min_column+1,
    max_col=max_column,
    min_row=min_row+3,
    max_row=max_row
)

categories = Reference(sheet,
    min_col=min_column,
    max_col=min_column,
    min_row=min_row+4,
    max_row=max_row
)

# Add a total row at the bottom of the pivot table
for i in range(min_column+1, max_column+1):
    letter = get_column_letter(i)
    sheet[f"{letter}{max_row+1}"] = f'=SUM({letter}{min_row+2}:{letter}{max_row})'
    sheet[f'{letter}{max_row+1}'].number_format = '"R"#,##0'


# Add the data and categories to the bar chart
barchart.add_data(data, titles_from_data=True)
barchart.set_categories(categories)
barchart.x_axis.delete = False
barchart.y_axis.delete = False

sheet.add_chart(barchart, "B10")

barchart.title = "Sales by Product line"
barchart.title.overlay = False
barchart.y_axis.title = "Total Sales (R)"
barchart.y_axis.title.overlay = False
barchart.style = 10
barchart.legend.overlay = False

output_path = os.path.join(application_path, f'report_{month}.xlsx')
wb.save(output_path)
os.startfile(output_path)